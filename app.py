
import streamlit as st
import pandas as pd
import pdfplumber
from io import BytesIO
import base64
from datetime import datetime
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter

# Initialisation de la session state
if 'df_final' not in st.session_state:
    st.session_state.df_final = None
if 'files_processed' not in st.session_state:
    st.session_state.files_processed = []

def extract_data_from_pdf(file):
    """
    Extracts transaction data from a single PDF file with robust date checking.
    """
    transactions = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            # Use extract_table with table settings for better parsing
            table = page.extract_table()
            if table:
                for row in table:
                    # Ensure row is a list and has enough columns
                    if not isinstance(row, list) or len(row) < 5:
                        continue

                    # Robustly check if the first column is a valid date
                    try:
                        pd.to_datetime(row[0], format='%d/%m/%Y', errors='raise')
                        date = row[0]
                    except (ValueError, TypeError, IndexError):
                        # This row does not start with a valid date, so we skip it
                        continue
                    
                    # If we have a valid date, we can assume it's a transaction row
                    try:
                        libelle = row[2]
                        debit_str = row[3]
                        credit_str = row[4]

                        amount = 0.0
                        if debit_str and debit_str.strip():
                            amount = -float(debit_str.replace('.', '').replace(',', '.'))
                        elif credit_str and credit_str.strip():
                            amount = float(credit_str.replace('.', '').replace(',', '.'))

                        # Only add rows with a non-zero amount
                        if amount != 0.0:
                            transactions.append([date, libelle, amount])
                    except (ValueError, IndexError):
                        # Skip rows that look like transactions but have parsing issues
                        continue
    return transactions

def to_excel(df):
    """
    Converts a DataFrame to an Excel file in memory, with auto-width and currency formatting.
    """
    output = BytesIO()
    df_copy = df.copy()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_copy.to_excel(writer, index=False, sheet_name='Transactions')

        workbook = writer.book
        worksheet = writer.sheets['Transactions']

        # Define and add the currency style to the workbook
        currency_style = NamedStyle(name='currency_style', number_format='#,##0.00 €')
        if currency_style.name not in workbook.style_names:
            workbook.add_named_style(currency_style)

        # Define and add the date style to the workbook
        date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
        if date_style.name not in workbook.style_names:
            workbook.add_named_style(date_style)

        # Auto-adjust columns and apply formats
        for idx, col in enumerate(worksheet.columns, 1):
            column_letter = get_column_letter(idx)
            max_length = 0
            
            # Find the maximum length of a cell in the column
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Adjust the column width
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column_letter].width = adjusted_width

            # Apply currency format to 'debit' and 'credit' columns
            header_cell = worksheet.cell(row=1, column=idx)
            if header_cell.value == 'date':
                for cell in col[1:]:
                    cell.style = date_style
            elif header_cell.value in ('debit', 'credit'):
                # Apply the style to all cells in the column except the header
                for cell in col[1:]:
                    cell.style = currency_style

    processed_data = output.getvalue()
    return processed_data

def calculate_statistics(df):
    """Calcule les statistiques sur les transactions"""
    if df is None or df.empty:
        return None
    
    total_debits = df['debit'].sum()
    total_credits = df['credit'].sum()
    solde_total = total_credits + total_debits  # débits sont négatifs
    nb_transactions = len(df)
    date_min = df['date'].min()
    date_max = df['date'].max()
    
    return {
        'solde_total': solde_total,
        'total_debits': abs(total_debits),
        'total_credits': total_credits,
        'nb_transactions': nb_transactions,
        'date_min': date_min,
        'date_max': date_max
    }

def filter_dataframe(df, search_term, date_range, amount_range, type_filter):
    """Filtre le DataFrame selon les critères"""
    filtered_df = df.copy()
    
    # Filtre par recherche de libellé
    if search_term:
        filtered_df = filtered_df[filtered_df['libelle'].str.contains(search_term, case=False, na=False)]
    
    # Filtre par date
    if date_range:
        if isinstance(date_range, tuple) and len(date_range) == 2:
            filtered_df = filtered_df[
                (filtered_df['date'] >= pd.Timestamp(date_range[0])) & 
                (filtered_df['date'] <= pd.Timestamp(date_range[1]))
            ]
        elif isinstance(date_range, (list, tuple)) and len(date_range) == 1:
            filtered_df = filtered_df[filtered_df['date'].dt.date == date_range[0]]
    
    # Filtre par montant
    if amount_range:
        montants = filtered_df['debit'] + filtered_df['credit']
        filtered_df = filtered_df[
            (montants >= amount_range[0]) & 
            (montants <= amount_range[1])
        ]
    
    # Filtre par type (débit/crédit)
    if type_filter == 'Débits uniquement':
        filtered_df = filtered_df[filtered_df['debit'] != 0]
    elif type_filter == 'Crédits uniquement':
        filtered_df = filtered_df[filtered_df['credit'] != 0]
    
    return filtered_df

st.set_page_config(
    layout="wide",
    page_title="Extracteur CIC",
    page_icon="📄"
)

st.title("📄 Extracteur de Relevés Bancaires CIC")

st.info("Cette application Streamlit permet d'extraire les transactions de relevés bancaires PDF de la banque CIC et de les exporter dans un fichier Excel propre et formaté.")

# Section principale avec colonnes pour l'upload et les actions
col1, col2 = st.columns([3, 1])

with col1:
    uploaded_files = st.file_uploader(
        "Sélectionnez vos fichiers PDF", 
        type="pdf", 
        accept_multiple_files=True,
        help="Vous pouvez sélectionner un ou plusieurs fichiers PDF de relevés bancaires CIC"
    )

with col2:
    if st.session_state.df_final is not None:
        if st.button("🗑️ Effacer", type="secondary", use_container_width=True):
            st.session_state.df_final = None
            st.session_state.files_processed = []
            st.rerun()

# Validation des fichiers
if uploaded_files:
    # Vérifier que ce sont bien des PDF
    invalid_files = [f.name for f in uploaded_files if not f.name.lower().endswith('.pdf')]
    if invalid_files:
        st.error(f"⚠️ Les fichiers suivants ne sont pas des PDF : {', '.join(invalid_files)}")
    
    # Traitement des fichiers
    if st.button("🚀 Extraire et Traiter les Données", type="primary", use_container_width=True):
        all_transactions = []
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        total_files = len(uploaded_files)
        
        for idx, uploaded_file in enumerate(uploaded_files):
            try:
                status_text.text(f"📄 Traitement de {uploaded_file.name} ({idx + 1}/{total_files})...")
                transactions = extract_data_from_pdf(uploaded_file)
                all_transactions.extend(transactions)
                progress_bar.progress((idx + 1) / total_files)
            except Exception as e:
                st.error(f"❌ Erreur lors du traitement du fichier {uploaded_file.name}: {e}")
        
        status_text.empty()
        progress_bar.empty()

        if all_transactions:
            # Création du DataFrame
            df = pd.DataFrame(all_transactions, columns=['date_str', 'libelle', 'montant'])
            
            df_final = pd.DataFrame({
                'date': pd.to_datetime(df['date_str'], format='%d/%m/%Y'),
                'libelle': df['libelle'],
                'debit': df['montant'].apply(lambda x: x if x < 0 else 0),
                'credit': df['montant'].apply(lambda x: x if x > 0 else 0)
            })
            
            df_final = df_final.sort_values(by='date').reset_index(drop=True)
            
            # Sauvegarder dans session_state
            st.session_state.df_final = df_final
            st.session_state.files_processed = [f.name for f in uploaded_files]
            
            st.success(f"✅ {len(df_final)} transactions extraites avec succès depuis {total_files} fichier(s) !")
            st.balloons()  # Animation de célébration
            st.rerun()
        else:
            st.warning("⚠️ Aucune transaction n'a pu être extraite des fichiers fournis. Vérifiez que les fichiers sont bien des relevés bancaires CIC valides.")

# Affichage des résultats si des données sont disponibles
if st.session_state.df_final is not None:
    df_final = st.session_state.df_final
    
    # Statistiques
    stats = calculate_statistics(df_final)
    
    if stats:
        st.markdown("---")
        st.subheader("📊 Statistiques")
        
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            st.metric("💳 Solde Total", f"{stats['solde_total']:,.2f} €")
        
        with col2:
            st.metric("📉 Total Débits", f"{stats['total_debits']:,.2f} €")
        
        with col3:
            st.metric("📈 Total Crédits", f"{stats['total_credits']:,.2f} €")
        
        with col4:
            st.metric("🔢 Transactions", f"{stats['nb_transactions']:,}")
        
        with col5:
            st.metric("📅 Période", 
                     f"{stats['date_min'].strftime('%d/%m/%Y')} → {stats['date_max'].strftime('%d/%m/%Y')}")
    
    # Filtres et recherche
    st.markdown("---")
    st.subheader("🔍 Filtres et Recherche")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        search_term = st.text_input("🔎 Rechercher dans les libellés", "")
    
    with col2:
        date_range = st.date_input(
            "📅 Filtrer par date",
            value=(df_final['date'].min(), df_final['date'].max()),
            min_value=df_final['date'].min(),
            max_value=df_final['date'].max()
        )
    
    with col3:
        min_amount = float(df_final[['debit', 'credit']].sum(axis=1).min())
        max_amount = float(df_final[['debit', 'credit']].sum(axis=1).max())
        amount_range = st.slider(
            "💰 Filtrer par montant (€)",
            min_value=float(min_amount),
            max_value=float(max_amount),
            value=(float(min_amount), float(max_amount)),
            step=0.01
        )
    
    with col4:
        type_filter = st.selectbox(
            "📋 Type de transaction",
            ["Toutes", "Débits uniquement", "Crédits uniquement"]
        )
    
    # Appliquer les filtres
    date_filter_active = isinstance(date_range, (tuple, list)) and len(date_range) == 2
    amount_filter_active = amount_range[0] != min_amount or amount_range[1] != max_amount
    
    df_filtered = filter_dataframe(
        df_final, 
        search_term, 
        date_range if date_filter_active else None, 
        amount_range if amount_filter_active else None,
        type_filter
    )
    
    # Affichage du tableau filtré
    st.markdown("---")
    st.subheader(f"📋 Transactions ({len(df_filtered)} sur {len(df_final)})")
    
    # Style du tableau avec alternance de couleurs
    def highlight_rows(row):
        """Applique une couleur de fond alternée pour les lignes"""
        return ['background-color: #f0f2f6' if row.name % 2 == 0 else '' for _ in row]
    
    styled_df = df_filtered.style.format({
        "debit": "{:.2f} €",
        "credit": "{:.2f} €",
        "date": "{:%d/%m/%Y}"
    }).apply(highlight_rows, axis=1)
    
    st.dataframe(
        styled_df,
        use_container_width=True,
        height=400
    )
    
    # Export Excel
    st.markdown("---")
    col1, col2 = st.columns([1, 4])
    
    with col1:
        date_str = datetime.now().strftime("%Y-%m-%d")
        excel_data = to_excel(df_filtered)
        st.download_button(
            label="📥 Télécharger Excel",
            data=excel_data,
            file_name=f"transactions_cic_{date_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    with col2:
        if st.session_state.files_processed:
            st.caption(f"📁 Fichiers traités : {', '.join(st.session_state.files_processed)}")

elif not uploaded_files:
    st.warning("⚠️ Veuillez charger au moins un fichier PDF pour commencer.")

# Footer with Logo, Version, and Copyright
import base64

def get_base64_of_bin_file(bin_file):
    with open(bin_file, 'rb') as f:
        data = f.read()
    return base64.b64encode(data).decode()

def get_img_with_href(local_img_path):
    img_format = local_img_path.split('.')[-1]
    bin_str = get_base64_of_bin_file(local_img_path)
    html_code = f'<img src="data:image/{img_format};base64,{bin_str}" width="70" />'
    return html_code

logo_html = get_img_with_href('assets/mon_logo.png')

footer_css = """
<style>
    /* Force le footer à rester fixe même lors des changements de contenu */
    .stApp {
        position: relative;
    }
    
    .footer-container {
        position: fixed !important;
        right: 10px !important;
        bottom: 10px !important;
        text-align: right;
        color: grey;
        z-index: 9999 !important;
        pointer-events: none;
        background: transparent;
        transform: translateZ(0);
        will-change: transform;
        margin: 0 !important;
        padding: 0 !important;
    }
    
    .footer-container img, 
    .footer-container p {
        pointer-events: auto;
        margin: 0;
        padding: 0;
    }
    
    /* Empêche tout élément Streamlit de modifier la position */
    [data-testid="stAppViewContainer"] {
        position: relative !important;
    }
</style>
"""

footer_html = f"""
<div class="footer-container">
    {logo_html}
    <p>Version 05.11.25<br>
    &copy; 2025 - Tous droits réservés</p>
</div>
"""

# Utiliser un conteneur vide pour maintenir le footer
footer_placeholder = st.container()
with footer_placeholder:
    st.markdown(footer_css, unsafe_allow_html=True)
    st.markdown(footer_html, unsafe_allow_html=True)
